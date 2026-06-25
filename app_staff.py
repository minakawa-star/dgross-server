import os
import jwt
import datetime
import requests
import calendar
import bcrypt
from flask import jsonify, request, send_file
from supabase import create_client
from openpyxl import load_workbook
from functools import wraps

SUPABASE_STAFF_URL = os.environ.get("SUPABASE_STAFF_URL")
SUPABASE_STAFF_KEY = os.environ.get("SUPABASE_STAFF_KEY")
JWT_SECRET = os.environ.get("JWT_SECRET")

supabase_staff = create_client(SUPABASE_STAFF_URL, SUPABASE_STAFF_KEY)

MASTER_PATH = os.path.join(os.path.dirname(__file__), "スタッフマスター.xlsx")

B_TO_D = {
    "B0000106": "D0000295",
    "B0000107": "D0000326",
    "D0001318": "B0000095"
}

RATE_TABLE = {
    5: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    4: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
    3: {22: (2.30, 2.00), 21: (2.30, 2.00), 20: (2.45, 2.10),
        19: (2.45, 2.10), 18: (2.45, 2.10), 17: (2.60, 2.20),
        16: (2.60, 2.20), 15: (2.60, 2.20), 14: (2.60, 2.20),
        13: (2.75, 2.30), 12: (2.75, 2.30), 11: (2.75, 2.30),
        10: (2.75, 2.30), 9: (2.75, 2.30), 8: (2.95, 2.57),
        7: (2.95, 2.57), 6: (2.95, 2.57), 5: (2.95, 2.57)},
}

# インセンティブ用：休み日数→売上比重テーブル（週ラベル付き）
INCENTIVE_RATE_TABLE = [
    (1, 2.30, "週5"),   # 休み0-1日 → 週5
    (4, 2.45, "週4"),   # 休み2-4日 → 週4
    (8, 2.60, "週3"),   # 休み5-8日 → 週3
    (12, 2.75, "週2"),  # 休み9-12日 → 週2
]


def load_staff_master():
    wb = load_workbook(MASTER_PATH, data_only=True)
    ws1 = wb["スタッフマスター"]
    ws2 = wb["時給マスター"]

    master = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        staff_id, name, site, rank = row[0], row[1], row[2], row[3]
        if not staff_id or not rank:
            continue
        sid = str(staff_id).strip()
        sid = B_TO_D.get(sid, sid)
        master[sid] = {
            "name": name, "site": site, "rank": rank,
            "hourly_wage": 0, "mgmt_fee": 0,
            "work_pattern": 5, "monthly_salary": None
        }

    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        sid = str(row[0]).strip()
        sid = B_TO_D.get(sid, sid)
        wage = row[2]
        note = str(row[3]) if row[3] else ""
        if sid not in master:
            continue
        if "月給" in note:
            master[sid]["monthly_salary"] = int(str(wage).replace(",", "").strip()) if wage else 0
        else:
            master[sid]["hourly_wage"] = int(str(wage).replace(",", "").strip()) if wage else 0
        master[sid]["mgmt_fee"] = 3030 if "管理料" in note else 0

    return master


def jwt_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if not token:
            return jsonify({"error": "認証が必要です"}), 401
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            request.staff_id = payload.get("staff_id")
            request.role = payload.get("role")
        except jwt.ExpiredSignatureError:
            return jsonify({"error": "トークンが期限切れです"}), 401
        except jwt.InvalidTokenError:
            return jsonify({"error": "無効なトークンです"}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get("Authorization", "").replace("Bearer ", "")
        if not token:
            token = request.args.get("token", "")  # window.open等、ヘッダーを送れない場合のフォールバック
        if not token:
            return jsonify({"error": "認証が必要です"}), 401
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=["HS256"])
            if payload.get("role") != "admin":
                return jsonify({"error": "管理者権限が必要です"}), 403
            request.staff_id = payload.get("staff_id")
            request.role = payload.get("role")
        except jwt.InvalidTokenError:
            return jsonify({"error": "無効なトークンです"}), 401
        return f(*args, **kwargs)
    return decorated

def calc_auto_fb(master, apo_rows, att_rows, target_month, holidays_set, auto_settings, penalty_adjustments=None):
    """
    FB自動集計（行動量ボーナス・勤怠ペナルティ）
    penalty_adjustments: { staff_id: 補正後の最終ペナルティ金額 }（指定があれば自動計算結果を上書き）
    戻り値:
      breakdown: { staff_id: [ {name, category, amount}, ... ] }
      totals:    { staff_id: 合計金額 }
    """
    breakdown = {}
    totals = {}
    penalty_adjustments = penalty_adjustments or {}

    if not auto_settings:
        return breakdown, totals

    threshold = auto_settings.get("behavior_bonus_threshold", 45000)
    bonus_w5 = auto_settings.get("behavior_bonus_w5", 50000)
    bonus_w4 = auto_settings.get("behavior_bonus_w4", 30000)
    bonus_w3 = auto_settings.get("behavior_bonus_w3", 10000)
    penalty_per = auto_settings.get("penalty_per_count", 20000)

    # 営業日数を取得
    y, m, _ = map(int, target_month.split("-"))
    business_days = get_business_days(y, m, holidays_set)

    # 出勤日数・売上・ペナルティを集計
    work_days_map = {}
    penalty_map = {}
    sales_map = {}

    for row in att_rows:
        sid = B_TO_D.get(row["staff_id"], row["staff_id"])
        if sid not in master:
            continue

        # 出勤日数
        if (row.get("work_hours") or 0) > 0:
            work_days_map[sid] = work_days_map.get(sid, 0) + 1

        # 勤怠ペナルティ
        absence = str(row.get("absence_status") or "").strip()
        late = str(row.get("late") or "").strip()
        early = str(row.get("early_leave") or "").strip()

        count = 0
        if absence == "1":
            count += 1
        if late == "0":
            count += 1
        if early == "0":
            count += 1

        if count > 0:
            penalty_map[sid] = penalty_map.get(sid, 0) + count

    # 売上集計
    for row in apo_rows:
        sid = B_TO_D.get(row["staff_id"], row["staff_id"])
        if sid not in master:
            continue
        cancel = str(row.get("cancel_date") or "")
        if not cancel or cancel == "None":
            sales_map[sid] = sales_map.get(sid, 0) + row.get("amount", 0)

    # 行動量ボーナス判定
    for sid in master:
        work_days = work_days_map.get(sid, 0)
        sales = sales_map.get(sid, 0)
        rest_days = business_days - work_days

        # 週区分判定
        if rest_days <= 1:
            bonus = bonus_w5
            week_label = "週5"
        elif rest_days <= 4:
            bonus = bonus_w4
            week_label = "週4"
        elif rest_days <= 8:
            bonus = bonus_w3
            week_label = "週3"
        else:
            continue  # 対象外

        if work_days == 0:
            continue

        # 1日平均売上チェック
        daily_avg = sales / work_days
        if daily_avg < threshold:
            continue

        breakdown.setdefault(sid, []).append({
            "name": f"行動量ボーナス（{week_label}）",
            "category": "行動量ボーナス",
            "amount": bonus
        })
        totals[sid] = totals.get(sid, 0) + bonus

    # 勤怠ペナルティ（補正がある場合は補正額で上書き）
    all_penalty_staff = set(penalty_map.keys()) | set(penalty_adjustments.keys())
    for sid in all_penalty_staff:
        if sid not in master:
            continue
        if sid in penalty_adjustments:
            amount = penalty_adjustments[sid]
            count = penalty_map.get(sid, 0)
            breakdown.setdefault(sid, []).append({
                "name": f"勤怠ペナルティ（{count}回・手動補正済）",
                "category": "勤怠ペナルティ",
                "amount": amount
            })
        else:
            count = penalty_map[sid]
            amount = count * (-penalty_per)
            breakdown.setdefault(sid, []).append({
                "name": f"勤怠ペナルティ（{count}回）",
                "category": "勤怠ペナルティ",
                "amount": amount
            })
        totals[sid] = totals.get(sid, 0) + amount

    return breakdown, totals
def calc_campaign_fb(apo_rows, campaigns, bulk_amounts=None):
    """
    FBキャンペーンの集計。
    戻り値:
      breakdown: { staff_id: [ {name, category, amount, count?}, ... ] }
      totals:    { staff_id: 合計金額 }
    """
    breakdown = {}
    totals = {}
    bulk_map = {}
    if bulk_amounts:
        for b in bulk_amounts:
            bulk_map.setdefault(b["campaign_id"], []).append(b)

    for c in campaigns:
        target_types = c.get("target_types") or []
        exclude_resend = c.get("exclude_resend", True)
        target_staff_ids = c.get("target_staff_ids") or []
        start = c.get("start_date")
        end = c.get("end_date")
        calc_type = c.get("calc_type")
        amount = c.get("amount", 0)

        if calc_type == "bulk":
            for b in bulk_map.get(c["id"], []):
                sid = B_TO_D.get(b["staff_id"], b["staff_id"])
                breakdown.setdefault(sid, []).append({
                    "name": c.get("name"),
                    "category": c.get("category"),
                    "amount": b["amount"]
                })
                totals[sid] = totals.get(sid, 0) + b["amount"]
        elif calc_type == "fixed":
            for sid in target_staff_ids:
                sid_m = B_TO_D.get(sid, sid)
                breakdown.setdefault(sid_m, []).append({
                    "name": c.get("name"),
                    "category": c.get("category"),
                    "amount": amount
                })
                totals[sid_m] = totals.get(sid_m, 0) + amount
        else:  # per_unit / 指定案件(project_id絞り込み対応)
            target_project_ids = c.get("target_project_ids") or []
            counts = {}
            for row in apo_rows:
                acq = row.get("appointment_date")
                if not acq or not start or not end:
                    continue
                if not (start <= acq <= end):
                    continue
                if target_types and row.get("apo_type") not in target_types:
                    continue
                if target_project_ids and row.get("project_id") not in target_project_ids:
                    continue
                if exclude_resend and row.get("resend_status") == "再送":
                    continue
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if target_staff_ids and sid not in target_staff_ids:
                    continue
                counts[sid] = counts.get(sid, 0) + 1

            for sid, cnt in counts.items():
                amt = cnt * amount
                breakdown.setdefault(sid, []).append({
                    "name": c.get("name"),
                    "category": c.get("category"),
                    "amount": amt,
                    "count": cnt
                })
                totals[sid] = totals.get(sid, 0) + amt

    return breakdown, totals


def shift_month(target_month_str, delta):
    """target_month('YYYY-MM-01')をdeltaヶ月シフトした 'YYYY-MM-01' を返す"""
    y, m, _ = map(int, target_month_str.split("-"))
    total = (y * 12 + (m - 1)) + delta
    ny = total // 12
    nm = total % 12 + 1
    return f"{ny:04d}-{nm:02d}-01"


def get_business_days(year, month, holidays_set):
    """その月の平日数(土日・祝日を除く)を計算"""
    days_in_month = calendar.monthrange(year, month)[1]
    count = 0
    for day in range(1, days_in_month + 1):
        d = datetime.date(year, month, day)
        if d.weekday() < 5:  # Mon-Fri
            if d.isoformat() not in holidays_set:
                count += 1
    return count


def register_staff_routes(app):

    @app.route("/health_staff")
    def health_staff():
        return jsonify({"status": "ok", "service": "staff-dashboard"})

    @app.route("/staff/login", methods=["POST"])
    def staff_login():
        try:
            data = request.get_json()
            login_id = data.get("login_id", "").strip()
            password = data.get("password", "").strip()
            if not login_id or not password:
                return jsonify({"error": "IDとパスワードを入力してください"}), 400
            res = supabase_staff.table("staff_master")\
                .select("*").eq("login_id", login_id).execute()
            if not res.data:
                return jsonify({"error": "IDまたはパスワードが間違っています"}), 401
            staff = res.data[0]
            if not bcrypt.checkpw(password.encode(), staff["password_hash"].encode()):
                return jsonify({"error": "IDまたはパスワードが間違っています"}), 401
            token = jwt.encode({
                "staff_id": staff["staff_id"],
                "role": staff["role"],
                "exp": datetime.datetime.utcnow() + datetime.timedelta(days=30)
            }, JWT_SECRET, algorithm="HS256")
            return jsonify({
                "status": "ok",
                "token": token,
                "staff_id": staff["staff_id"],
                "role": staff["role"],
                "name": staff["staff_name"],
                "password_changed": staff.get("password_changed", False)
            })
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/register", methods=["POST"])
    @admin_required
    def staff_register():
        try:
            data = request.get_json()
            staff_id = data.get("staff_id", "").strip()
            staff_name = data.get("staff_name", "").strip()
            role = data.get("role", "staff")
            if not all([staff_id, staff_name]):
                return jsonify({"error": "必須項目が不足しています"}), 400
            default_password = "Dghojin2026"
            password_hash = bcrypt.hashpw(default_password.encode(), bcrypt.gensalt()).decode()
            supabase_staff.table("staff_master").upsert({
                "staff_id": staff_id,
                "staff_name": staff_name,
                "login_id": staff_id,
                "password_hash": password_hash,
                "role": role,
                "password_changed": False
            }).execute()
            return jsonify({"status": "ok", "staff_id": staff_id})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/me")
    @jwt_required
    def staff_me():
        try:
            res = supabase_staff.table("staff_master")\
                .select("staff_id,staff_name,role")\
                .eq("staff_id", request.staff_id).execute()
            if not res.data:
                return jsonify({"error": "スタッフが見つかりません"}), 404
            return jsonify({"status": "ok", "data": res.data[0]})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/staff/change_password", methods=["POST"])
    @jwt_required
    def staff_change_password():
        """本人によるパスワード変更（初回ログイン時の強制変更にも使用）"""
        try:
            data = request.get_json()
            new_password = data.get("new_password", "").strip()
            if not new_password or len(new_password) < 8:
                return jsonify({"error": "パスワードは8文字以上で入力してください"}), 400
            password_hash = bcrypt.hashpw(new_password.encode(), bcrypt.gensalt()).decode()
            supabase_staff.table("staff_master").update({
                "password_hash": password_hash,
                "password_changed": True
            }).eq("staff_id", request.staff_id).execute()
            return jsonify({"status": "ok"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/penalty_adjustments", methods=["GET", "POST"])
    @admin_required
    def staff_penalty_adjustments():
        """勤怠ペナルティの手動補正（合計値で登録/取得）"""
        try:
            if request.method == "GET":
                month = request.args.get("month")
                if not month:
                    return jsonify({"error": "monthパラメータが必要です"}), 400
                target_month = month + "-01"
                res = supabase_staff.table("fb_penalty_adjustments")\
                    .select("*").eq("target_month", target_month).execute()
                return jsonify({"status": "ok", "data": res.data})
            else:
                data = request.get_json()
                staff_id = data.get("staff_id", "").strip()
                month = data.get("month", "").strip()
                amount = data.get("amount")
                reason = data.get("reason", "").strip()
                if not staff_id or not month or amount is None:
                    return jsonify({"error": "staff_id, month, amountが必要です"}), 400
                target_month = month + "-01"
                supabase_staff.table("fb_penalty_adjustments").upsert({
                    "staff_id": staff_id,
                    "target_month": target_month,
                    "amount": int(amount),
                    "reason": reason
                }, on_conflict="staff_id,target_month").execute()
                return jsonify({"status": "ok"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/penalty_adjustments/<staff_id>", methods=["DELETE"])
    @admin_required
    def staff_penalty_adjustments_delete(staff_id):
        """補正の削除（自動計算に戻す）"""
        try:
            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthパラメータが必要です"}), 400
            target_month = month + "-01"
            supabase_staff.table("fb_penalty_adjustments")\
                .delete().eq("staff_id", staff_id).eq("target_month", target_month).execute()
            return jsonify({"status": "ok"})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/staff/admin_reset_password", methods=["POST"])
    @admin_required
    def staff_admin_reset_password():
        """管理者による特定スタッフのパスワードリセット（初期パスワードに戻し、未変更フラグも戻す）"""
        try:
            data = request.get_json()
            staff_id = data.get("staff_id", "").strip()
            if not staff_id:
                return jsonify({"error": "staff_idが必要です"}), 400
            default_password = "Dghojin2026"
            password_hash = bcrypt.hashpw(default_password.encode(), bcrypt.gensalt()).decode()
            supabase_staff.table("staff_master").update({
                "password_hash": password_hash,
                "password_changed": False
            }).eq("staff_id", staff_id).execute()
            return jsonify({"status": "ok"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/debug_master")
    def debug_master():
        try:
            master = load_staff_master()
            targets = ["B0000002", "B0000032", "D0000295", "D0000326", "D0001221", "D0001316"]
            result = {k: v for k, v in master.items() if k in targets}
            return jsonify(result)
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/target", methods=["GET", "POST"])
    @jwt_required
    def staff_target():
        if request.method == "GET":
            try:
                staff_id = request.args.get("staff_id")
                month = request.args.get("month")
                if not staff_id or not month:
                    return jsonify({"error": "staff_id, monthが必要です"}), 400

                if request.role != "admin" and staff_id != request.staff_id:
                    return jsonify({"error": "権限がありません"}), 403

                target_month = month + "-01"
                res = supabase_staff.table("monthly_targets")\
                    .select("*").eq("staff_id", staff_id).eq("target_month", target_month).execute()

                if res.data:
                    return jsonify({"status": "ok", "data": res.data[0]})
                else:
                    return jsonify({"status": "ok", "data": {
                        "staff_id": staff_id, "target_month": target_month,
                        "planned_work_days": 0, "is_confirmed": False, "confirmed_work_days": None
                    }})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

        else:
            try:
                data = request.get_json()
                staff_id = data.get("staff_id")
                month = data.get("month")
                planned_work_days = data.get("planned_work_days")

                if not staff_id or not month:
                    return jsonify({"error": "staff_id, monthが必要です"}), 400

                if request.role != "admin" and staff_id != request.staff_id:
                    return jsonify({"error": "権限がありません"}), 403

                target_month = month + "-01"

                if request.role != "admin":
                    existing = supabase_staff.table("monthly_targets")\
                        .select("is_confirmed").eq("staff_id", staff_id).eq("target_month", target_month).execute()
                    if existing.data and existing.data[0]["is_confirmed"]:
                        return jsonify({"error": "確定済みのため編集できません"}), 403

                supabase_staff.table("monthly_targets").upsert({
                    "staff_id": staff_id,
                    "target_month": target_month,
                    "planned_work_days": int(planned_work_days)
                }, on_conflict="staff_id,target_month").execute()

                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/target/bulk", methods=["POST"])
    @admin_required
    def staff_target_bulk():
        """出勤予定数の一括登録（未入力の人だけ、または全員に同じ値を反映）"""
        try:
            data = request.get_json()
            month = data.get("month")
            value = data.get("planned_work_days")
            staff_ids = data.get("staff_ids")  # 指定があればこの人たちだけ、なければ全員
            only_empty = data.get("only_empty", True)  # Trueなら未入力の人だけ上書き

            if not month or value is None:
                return jsonify({"error": "month, planned_work_daysが必要です"}), 400

            target_month = month + "-01"
            master = load_staff_master()
            target_staff_ids = staff_ids if staff_ids else list(master.keys())

            existing_res = supabase_staff.table("monthly_targets")\
                .select("*").eq("target_month", target_month).execute()
            existing_map = {e["staff_id"]: e for e in existing_res.data}

            updated = []
            skipped = []
            for sid in target_staff_ids:
                existing = existing_map.get(sid)
                if existing and existing.get("is_confirmed"):
                    skipped.append(sid)
                    continue
                if only_empty and existing and existing.get("planned_work_days"):
                    skipped.append(sid)
                    continue
                supabase_staff.table("monthly_targets").upsert({
                    "staff_id": sid,
                    "target_month": target_month,
                    "planned_work_days": int(value)
                }, on_conflict="staff_id,target_month").execute()
                updated.append(sid)

            return jsonify({"status": "ok", "updated_count": len(updated), "skipped_count": len(skipped)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/month_status")
    @admin_required
    def month_status():
        """指定月の①目標値確定・②時給確定の状況を返す"""
        try:
            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400
            target_month = month + "-01"

            targets_res = supabase_staff.table("monthly_targets")\
                .select("is_confirmed").eq("target_month", target_month).execute()
            target_confirmed = bool(targets_res.data) and all(t.get("is_confirmed") for t in targets_res.data)

            wage_res = supabase_staff.table("wage_confirm_status")\
                .select("*").eq("target_month", target_month).execute()
            wage_confirmed = bool(wage_res.data) and wage_res.data[0].get("is_confirmed", False)

            return jsonify({"status": "ok", "target_confirmed": target_confirmed, "wage_confirmed": wage_confirmed})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/confirm_month", methods=["POST"])
    @admin_required
    def confirm_month():
        try:
            data = request.get_json()
            month = data.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400

            target_month = month + "-01"
            master = load_staff_master()

            att_res = supabase_staff.table("attendance")\
                .select("*").eq("target_month", target_month).execute()

            work_days_map = {}
            for row in att_res.data:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if (row.get("work_hours") or 0) > 0:
                    work_days_map[sid] = work_days_map.get(sid, 0) + 1

            for sid in master.keys():
                days = work_days_map.get(sid, 0)
                supabase_staff.table("monthly_targets").upsert({
                    "staff_id": sid,
                    "target_month": target_month,
                    "is_confirmed": True,
                    "confirmed_work_days": days
                }, on_conflict="staff_id,target_month").execute()

            return jsonify({"status": "ok", "message": f"{month}を確定しました"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/unconfirm_month", methods=["POST"])
    @admin_required
    def unconfirm_month():
        """目標値確定の解除。時給確定が済んでいる場合はエラーで止める"""
        try:
            data = request.get_json()
            month = data.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400
            target_month = month + "-01"

            wage_status_res = supabase_staff.table("wage_confirm_status")\
                .select("*").eq("target_month", target_month).execute()
            if wage_status_res.data and wage_status_res.data[0].get("is_confirmed"):
                return jsonify({"error": "時給確定が済んでいるため解除できません。先に時給確定を解除してください"}), 400

            supabase_staff.table("monthly_targets")\
                .update({"is_confirmed": False, "confirmed_work_days": None})\
                .eq("target_month", target_month).execute()

            return jsonify({"status": "ok", "message": f"{month}の目標値確定を解除しました"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/confirm_wage", methods=["POST"])
    @admin_required
    def confirm_wage():
        """確定時給を確定し、履歴テーブルに保存する。目標値確定が済んでいない月はエラー"""
        try:
            data = request.get_json()
            month = data.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400
            target_month = month + "-01"

            targets_res = supabase_staff.table("monthly_targets")\
                .select("*").eq("target_month", target_month).execute()
            if not targets_res.data or not any(t.get("is_confirmed") for t in targets_res.data):
                return jsonify({"error": "目標値確定が済んでいないため、時給確定できません"}), 400

            # 現在の/staff/summaryと同じロジックで確定時給を算出（直接呼び出し、HTTP往復を避ける）
            summary_res = build_staff_summary(month)
            if "error" in summary_res:
                return jsonify(summary_res), 500

            rows = []
            for r in summary_res["data"]:
                if r.get("confirmed_wage") is not None:
                    rows.append({
                        "staff_id": r["staff_id"],
                        "target_month": target_month,
                        "confirmed_wage": r["confirmed_wage"],
                        "wage_change_label": r["confirmed_wage_change"]
                    })
            if rows:
                supabase_staff.table("confirmed_wage_history")\
                    .upsert(rows, on_conflict="staff_id,target_month").execute()

            supabase_staff.table("wage_confirm_status").upsert({
                "target_month": target_month,
                "is_confirmed": True,
                "confirmed_at": datetime.datetime.utcnow().isoformat()
            }).execute()

            return jsonify({"status": "ok", "message": f"{month}の時給を確定しました", "count": len(rows)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/unconfirm_wage", methods=["POST"])
    @admin_required
    def unconfirm_wage():
        """時給確定の解除"""
        try:
            data = request.get_json()
            month = data.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400
            target_month = month + "-01"

            supabase_staff.table("wage_confirm_status")\
                .update({"is_confirmed": False}).eq("target_month", target_month).execute()
            supabase_staff.table("confirmed_wage_history")\
                .delete().eq("target_month", target_month).execute()

            return jsonify({"status": "ok", "message": f"{month}の時給確定を解除しました"})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # Excel出力（インセンティブ支給額・確定時給）
    # ============================================================
    @app.route("/staff/export/incentive_payout")
    @admin_required
    def export_incentive_payout():
        try:
            from openpyxl import Workbook
            import io

            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400

            summary_res = build_staff_summary(month)
            summary_data = summary_res.get("data", [])

            wb = Workbook()
            ws = wb.active
            ws.title = "インセンティブ支給額"
            ws.append(["社員番号", "スタッフ名", "支払金額"])
            for r in summary_data:
                if r.get("incentive_payout_status") == "ok" and r.get("incentive_payout", 0) > 0:
                    ws.append([r["staff_id"], r["name"], r["incentive_payout"]])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, download_name=f"インセンティブ支給額_{month}.xlsx",
                              as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/export/confirmed_wage")
    @admin_required
    def export_confirmed_wage():
        try:
            from openpyxl import Workbook
            import io

            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400

            target_month = month + "-01"
            prev_month = shift_month(target_month, -1)[:7]

            master = load_staff_master()

            # 次月時給＝前月分の確定時給履歴
            prev_history_res = supabase_staff.table("confirmed_wage_history")\
                .select("*").eq("target_month", prev_month + "-01").execute()
            prev_history_map = {h["staff_id"]: h for h in prev_history_res.data}

            # 翌々月時給＝当月の確定時給結果（履歴に保存済みであれば使用、なければ計算結果から取得）
            cur_history_res = supabase_staff.table("confirmed_wage_history")\
                .select("*").eq("target_month", target_month).execute()
            cur_history_map = {h["staff_id"]: h for h in cur_history_res.data}

            if not cur_history_map:
                summary_res = build_staff_summary(month)
                summary_data = summary_res.get("data", [])
                for r in summary_data:
                    if r.get("confirmed_wage") is not None:
                        cur_history_map[r["staff_id"]] = {
                            "confirmed_wage": r["confirmed_wage"],
                            "wage_change_label": r["confirmed_wage_change"]
                        }

            wb = Workbook()
            ws = wb.active
            ws.title = "確定時給"
            ws.append(["社員番号", "オペレーター名", "次月時給", "翌々月時給", "変給対象"])

            for sid, info in master.items():
                prev = prev_history_map.get(sid)
                cur = cur_history_map.get(sid)
                next_wage = prev["confirmed_wage"] if prev else None
                next_next_wage = cur["confirmed_wage"] if cur else None
                if next_wage is not None and next_next_wage is not None and next_wage != next_next_wage:
                    change_mark = "昇給" if next_next_wage > next_wage else "降給"
                else:
                    change_mark = ""
                if next_wage is None and next_next_wage is None:
                    continue
                ws.append([sid, info["name"], next_wage, next_next_wage, change_mark])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, download_name=f"確定時給_{month}.xlsx",
                              as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/export/fb_campaign_detail")
    @admin_required
    def export_fb_campaign_detail():
        """FBキャンペーン詳細リストのExcel出力（シート1：キャンペーン一覧、シート2：スタッフ別FB受取明細）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import io

            month = request.args.get("month")
            if not month:
                return jsonify({"error": "monthが必要です"}), 400

            campaigns_res = supabase_staff.table("fb_campaigns").select("*").execute()
            campaigns = campaigns_res.data

            summary_res = build_staff_summary(month)
            summary_data = summary_res.get("data", [])

            wb = Workbook()

            header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            header_fill_blue = PatternFill("solid", fgColor="005BAC")
            header_fill_green = PatternFill("solid", fgColor="375623")
            left = Alignment(horizontal="left", vertical="center")
            thin = Side(style="thin", color="CCCCCC")
            thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def style_header(ws, row, col_count, fill):
                for c in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

            def style_data_row(ws, row, col_count):
                for c in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=c)
                    cell.alignment = left
                    cell.border = thin_border

            # シート1：キャンペーン一覧
            ws1 = wb.active
            ws1.title = "キャンペーン一覧"
            headers1 = ["キャンペーン名", "カテゴリ", "計算方式", "開始日", "終了日", "金額", "対象種別", "対象スタッフ数"]
            for c, h in enumerate(headers1, 1):
                ws1.cell(row=1, column=c, value=h)
            style_header(ws1, 1, len(headers1), header_fill_blue)
            for w, col in zip([28, 18, 14, 12, 12, 14, 20, 14], ["A","B","C","D","E","F","G","H"]):
                ws1.column_dimensions[col].width = w

            calc_label = {"per_unit": "件数×単価", "fixed": "固定額", "bulk": "一括付与"}
            for row_idx, c in enumerate(campaigns, 2):
                types_str = "・".join(c.get("target_types") or []) or "全種別"
                staff_count = len(c.get("target_staff_ids") or [])
                staff_str = f"{staff_count}名指定" if staff_count > 0 else "全員"
                vals = [c.get("name",""), c.get("category",""),
                        calc_label.get(c.get("calc_type",""), c.get("calc_type","")),
                        c.get("start_date",""), c.get("end_date",""),
                        c.get("amount", 0), types_str, staff_str]
                for col, v in enumerate(vals, 1):
                    ws1.cell(row=row_idx, column=col, value=v)
                style_data_row(ws1, row_idx, len(headers1))

            # シート2：スタッフ別FB受取明細
            ws2 = wb.create_sheet(title="スタッフ別FB受取明細")
            headers2 = ["社員番号", "スタッフ名", "サイト", "キャンペーン名", "カテゴリ", "金額"]
            for c, h in enumerate(headers2, 1):
                ws2.cell(row=1, column=c, value=h)
            style_header(ws2, 1, len(headers2), header_fill_green)
            for w, col in zip([14, 22, 14, 28, 18, 14], ["A","B","C","D","E","F"]):
                ws2.column_dimensions[col].width = w

            row_idx = 2
            for staff in sorted(summary_data, key=lambda x: x.get("site","") + x.get("name","")):
                for item in (staff.get("fb_breakdown") or []):
                    vals2 = [staff.get("staff_id",""), staff.get("name",""), staff.get("site",""),
                             item.get("name",""), item.get("category",""), item.get("amount", 0)]
                    for col, v in enumerate(vals2, 1):
                        ws2.cell(row=row_idx, column=col, value=v)
                    style_data_row(ws2, row_idx, len(headers2))
                    row_idx += 1

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, download_name=f"FBキャンペーン詳細_{month}.xlsx",
                              as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # FBキャンペーン管理
    # ============================================================
    @app.route("/staff/fb_campaigns", methods=["GET", "POST"])
    @admin_required
    def fb_campaigns():
        if request.method == "GET":
            try:
                res = supabase_staff.table("fb_campaigns").select("*").order("start_date", desc=True).execute()
                return jsonify({"status": "ok", "data": res.data})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                required = ["name", "category", "calc_type", "start_date", "end_date", "amount"]
                for f in required:
                    if data.get(f) is None or data.get(f) == "":
                        return jsonify({"error": f"{f}は必須です"}), 400

                record = {
                    "name": data["name"],
                    "category": data["category"],
                    "calc_type": data["calc_type"],
                    "start_date": data["start_date"],
                    "end_date": data["end_date"],
                    "amount": int(data["amount"]),
                    "target_types": data.get("target_types") or [],
                    "exclude_resend": data.get("exclude_resend", True),
                    "target_staff_ids": data.get("target_staff_ids") or [],
                    "target_project_ids": data.get("target_project_ids") or []
                }
                res = supabase_staff.table("fb_campaigns").insert(record).execute()
                return jsonify({"status": "ok", "data": res.data[0]})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/fb_campaigns/<int:campaign_id>", methods=["PUT", "DELETE"])
    @admin_required
    def fb_campaign_detail(campaign_id):
        if request.method == "DELETE":
            try:
                supabase_staff.table("fb_campaigns").delete().eq("id", campaign_id).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                required = ["name", "category", "calc_type", "start_date", "end_date", "amount"]
                for f in required:
                    if data.get(f) is None or data.get(f) == "":
                        return jsonify({"error": f"{f}は必須です"}), 400

                record = {
                    "name": data["name"],
                    "category": data["category"],
                    "calc_type": data["calc_type"],
                    "start_date": data["start_date"],
                    "end_date": data["end_date"],
                    "amount": int(data["amount"]),
                    "target_types": data.get("target_types") or [],
                    "exclude_resend": data.get("exclude_resend", True),
                    "target_staff_ids": data.get("target_staff_ids") or [],
                    "target_project_ids": data.get("target_project_ids") or []
                }
                supabase_staff.table("fb_campaigns").update(record).eq("id", campaign_id).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # インセンティブ設定（対象時給範囲）
    # ============================================================
    @app.route("/staff/incentive_settings", methods=["GET", "POST"])
    @admin_required
    def incentive_settings():
        if request.method == "GET":
            try:
                res = supabase_staff.table("incentive_settings").select("*").eq("id", 1).execute()
                if res.data:
                    return jsonify({"status": "ok", "data": res.data[0]})
                else:
                    return jsonify({"status": "ok", "data": {"id": 1, "min_wage": 2100, "max_wage": 2500}})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                min_wage = int(data.get("min_wage"))
                max_wage = int(data.get("max_wage"))
                if min_wage > max_wage:
                    return jsonify({"error": "下限は上限以下にしてください"}), 400
                supabase_staff.table("incentive_settings").upsert({
                    "id": 1, "min_wage": min_wage, "max_wage": max_wage
                }).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/incentive_payout_rates", methods=["GET", "POST"])
    @admin_required
    def incentive_payout_rates():
        """インセンティブ支給率テーブル（時給帯×区分）の取得・更新。区分は週勤務 or 達成率のいずれかのモードで運用"""
        if request.method == "GET":
            try:
                rates_res = supabase_staff.table("incentive_payout_rates").select("*").execute()
                mode_res = supabase_staff.table("incentive_payout_mode").select("*").eq("id", 1).execute()
                mode = mode_res.data[0]["mode"] if mode_res.data else "week"
                return jsonify({"status": "ok", "data": rates_res.data, "mode": mode})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
        else:
            try:
                data = request.get_json()
                rates = data.get("rates", [])
                mode = data.get("mode", "week")

                supabase_staff.table("incentive_payout_mode").upsert({"id": 1, "mode": mode}).execute()

                # 既存を全削除してから入れ直す（空欄にした項目は対象外として除外されるため）
                supabase_staff.table("incentive_payout_rates").delete().neq("id", 0).execute()
                if rates:
                    rows = []
                    for r in rates:
                        rows.append({
                            "wage_band": int(r["wage_band"]),
                            "segment_label": r["segment_label"],
                            "segment_min": r.get("segment_min"),
                            "segment_max": r.get("segment_max"),
                            "rate": float(r["rate"])
                        })
                    supabase_staff.table("incentive_payout_rates").insert(rows).execute()
                return jsonify({"status": "ok"})
            except Exception as e:
                import traceback
                return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # インセンティブ用給与データ（管理本部CSV）アップロード
    # ============================================================
    @app.route("/staff/upload/incentive_payroll_json", methods=["POST"])
    @admin_required
    def upload_incentive_payroll_json():
        try:
            data = request.get_json()
            month = data.get("month")  # "YYYY-MM"
            records = data.get("records", [])
            if not month or not records:
                return jsonify({"error": "month, recordsが必要です"}), 400

            target_month = month + "-01"
            rows = []
            for r in records:
                sid = str(r.get("staff_id", "")).strip()
                if not sid:
                    continue
                sid = B_TO_D.get(sid, sid)
                rows.append({
                    "staff_id": sid,
                    "target_month": target_month,
                    "base_salary": r.get("base_salary") or 0,
                    "overtime_allowance": r.get("overtime_allowance") or 0,
                    "commute_allowance": r.get("commute_allowance") or 0
                })

            if rows:
                supabase_staff.table("incentive_payroll").upsert(rows, on_conflict="staff_id,target_month").execute()
            return jsonify({"status": "ok", "count": len(rows)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    # ============================================================
    # メイン集計
    # ============================================================
    def build_staff_summary(month):
        """月次サマリー計算の本体。staff_summaryエンドポイントとconfirm_wageの両方から呼ばれる（自己HTTP呼び出しを避けてメモリ消費を抑えるため）"""
        if not month:
            return {"error": "monthパラメータが必要です"}

        target_month = month + "-01"
        master = load_staff_master()

        apo_res = supabase_staff.table("appointments")\
            .select("*").eq("target_month", target_month).execute()
        apo_rows = apo_res.data

        att_res = supabase_staff.table("attendance")\
            .select("*").eq("target_month", target_month).execute()
        att_rows = att_res.data

        targets_res = supabase_staff.table("monthly_targets")\
            .select("*").eq("target_month", target_month).execute()
        targets_map = {t["staff_id"]: t for t in targets_res.data}

        campaigns_res = supabase_staff.table("fb_campaigns").select("*").execute()
        bulk_res = supabase_staff.table("fb_bulk_amounts").select("*").execute()
        campaign_breakdown, campaign_totals = calc_campaign_fb(apo_rows, campaigns_res.data, bulk_res.data)

        holidays_res = supabase_staff.table("holidays").select("holiday_date").execute()
        holidays_set = {h["holiday_date"] for h in holidays_res.data}

        # 自動集計FB（勤怠ペナルティの手動補正を反映）
        auto_settings_res = supabase_staff.table("fb_auto_settings").select("*").eq("id", 1).execute()
        auto_settings = auto_settings_res.data[0] if auto_settings_res.data else {}
        adj_res = supabase_staff.table("fb_penalty_adjustments")\
            .select("*").eq("target_month", target_month).execute()
        penalty_adjustments = {a["staff_id"]: a["amount"] for a in adj_res.data}
        auto_breakdown, auto_totals = calc_auto_fb(master, apo_rows, att_rows, target_month, holidays_set, auto_settings, penalty_adjustments)
        # ---- インセンティブ計算用の事前データ ----
        settings_res = supabase_staff.table("incentive_settings").select("*").eq("id", 1).execute()
        if settings_res.data:
            inc_min = settings_res.data[0]["min_wage"]
            inc_max = settings_res.data[0]["max_wage"]
        else:
            inc_min, inc_max = 2100, 2500

        prior_month = shift_month(target_month, -2)  # 2ヶ月前

        prior_att_res = supabase_staff.table("attendance")\
            .select("*").eq("target_month", prior_month).execute()
        prior_work_days_map = {}
        for row in prior_att_res.data:
            sid = B_TO_D.get(row["staff_id"], row["staff_id"])
            if (row.get("work_hours") or 0) > 0:
                prior_work_days_map[sid] = prior_work_days_map.get(sid, 0) + 1

        prior_payroll_res = supabase_staff.table("incentive_payroll")\
            .select("*").eq("target_month", prior_month).execute()
        prior_payroll_map = {}
        for row in prior_payroll_res.data:
            sid = B_TO_D.get(row["staff_id"], row["staff_id"])
            prior_payroll_map[sid] = {
                "base_salary": row.get("base_salary") or 0,
                "overtime_allowance": row.get("overtime_allowance") or 0,
                "commute_allowance": row.get("commute_allowance") or 0
            }

        holidays_res = supabase_staff.table("holidays").select("holiday_date").execute()
        holidays_set = {h["holiday_date"] for h in holidays_res.data}

        py, pm, _ = map(int, prior_month.split("-"))
        prior_business_days = get_business_days(py, pm, holidays_set)

        # 2ヶ月前の売上実績（アポ取得金額－キャンセル金額＋FB）を計算
        prior_apo_res = supabase_staff.table("appointments")\
            .select("*").eq("target_month", prior_month).execute()
        prior_apo_rows = prior_apo_res.data

        prior_sales_map = {}
        for row in prior_apo_rows:
            sid = B_TO_D.get(row["staff_id"], row["staff_id"])
            if sid not in prior_sales_map:
                prior_sales_map[sid] = {"apo_amount": 0, "cxl_amount": 0}
            cancel = str(row.get("cancel_date") or "")
            amount = row.get("achievement_amount", 0)
            if cancel and cancel not in ["None", ""]:
                prior_sales_map[sid]["cxl_amount"] += amount
            else:
                prior_sales_map[sid]["apo_amount"] += amount

        prior_campaigns_res = supabase_staff.table("fb_campaigns").select("*").execute()
        prior_campaigns = prior_campaigns_res.data
        prior_campaign_breakdown, prior_campaign_totals = calc_campaign_fb(prior_apo_rows, prior_campaigns, bulk_res.data)

        prior_auto_settings_res = supabase_staff.table("fb_auto_settings").select("*").eq("id", 1).execute()
        prior_auto_settings = prior_auto_settings_res.data[0] if prior_auto_settings_res.data else {}
        prior_adj_res = supabase_staff.table("fb_penalty_adjustments")\
            .select("*").eq("target_month", prior_month).execute()
        prior_penalty_adjustments = {a["staff_id"]: a["amount"] for a in prior_adj_res.data}
        prior_auto_breakdown, prior_auto_totals = calc_auto_fb(
            master, prior_apo_rows, prior_att_res.data, prior_month, holidays_set,
            prior_auto_settings, prior_penalty_adjustments
        )

        prior_fb_totals = {}
        for sid in set(list(prior_campaign_totals.keys()) + list(prior_auto_totals.keys())):
            prior_fb_totals[sid] = prior_campaign_totals.get(sid, 0) + prior_auto_totals.get(sid, 0)

        prior_sales_actual_map = {}
        for sid, vals in prior_sales_map.items():
            fb = prior_fb_totals.get(sid, 0)
            prior_sales_actual_map[sid] = vals["apo_amount"] - vals["cxl_amount"] + fb

        # 還元率テーブルの取得
        payout_rates_res = supabase_staff.table("incentive_payout_rates").select("*").execute()
        payout_rate_rows = payout_rates_res.data

        payout_mode_res = supabase_staff.table("incentive_payout_mode").select("*").eq("id", 1).execute()
        payout_mode = payout_mode_res.data[0]["mode"] if payout_mode_res.data else "week"

        # ---- 結果初期化 ----
        results = {}
        for sid, info in master.items():
            results[sid] = {
                "staff_id": sid,
                "name": info["name"],
                "site": info["site"],
                "rank": info["rank"],
                "apo_amount": 0,
                "cxl_amount": 0,
                "fb_amount": 0,
                "fb_breakdown": [],
                "sales": 0,
                "work_days": 0,
                "target_achieve": 0,
                "target_maintain": 0,
                "achieve_rate": None,
                "is_monthly": info["monthly_salary"] is not None,
                "hourly_wage": info["hourly_wage"],
                "monthly_salary": info["monthly_salary"],
                "planned_work_days": 0,
                "is_confirmed": False,
                "incentive_target": None,
                "incentive_status": "対象外",
                "incentive_detail": None
            }

        for row in apo_rows:
            sid = B_TO_D.get(row["staff_id"], row["staff_id"])
            if sid not in results:
                continue
            cancel = str(row.get("cancel_date") or "")
            if cancel and cancel not in ["None", ""]:
                results[sid]["cxl_amount"] += row.get("achievement_amount", 0)
            else:
                results[sid]["apo_amount"] += row.get("achievement_amount", 0)
                
        for row in att_rows:
            sid = B_TO_D.get(row["staff_id"], row["staff_id"])
            if sid not in results:
                continue
            if (row.get("work_hours") or 0) > 0:
                results[sid]["work_days"] += 1

        for sid, r in results.items():
            info = master[sid]

            # FB内訳：キャンペーン分
            for entry in campaign_breakdown.get(sid, []):
                r["fb_breakdown"].append(entry)

            campaign_total = campaign_totals.get(sid, 0)
            auto_total = auto_totals.get(sid, 0)
            r["fb_amount"] = campaign_total + auto_total

            # FB内訳にauto_fbを追加
            for entry in auto_breakdown.get(sid, []):
                r["fb_breakdown"].append(entry)

            r["sales"] = r["apo_amount"] - r["cxl_amount"] + r["fb_amount"]

            tgt = targets_map.get(sid)
            is_confirmed = tgt["is_confirmed"] if tgt else False
            r["is_confirmed"] = is_confirmed
            r["planned_work_days"] = tgt["planned_work_days"] if tgt else 0

            if is_confirmed:
                calc_days = tgt.get("confirmed_work_days")
                if calc_days is None:
                    calc_days = r["work_days"]
            else:
                calc_days = tgt["planned_work_days"] if (tgt and tgt["planned_work_days"] > 0) else r["work_days"]

            # ---- 通常の達成/維持目標 ----
            if info["monthly_salary"] is not None:
                base = info["monthly_salary"] * 1.15 + 20000
                r["target_achieve"] = int(base / 0.40)
                r["target_maintain"] = int(base / 0.45)
                if r["work_days"] == 0:
                    r["work_days"] = calc_days if calc_days > 0 else 22
            else:
                wage = info["hourly_wage"]
                mgmt = info["mgmt_fee"]
                pattern = info["work_pattern"]
                days = calc_days
                rate_row = RATE_TABLE.get(pattern, {}).get(days)
                if rate_row:
                    base = wage * 8 + 1000 + mgmt
                    r["target_achieve"] = int(base * days * rate_row[0])
                    r["target_maintain"] = int(base * days * rate_row[1])

            if r["target_achieve"] > 0:
                r["achieve_rate"] = round(r["sales"] / r["target_achieve"] * 100, 1)

            # ---- 確定時給（暫定ロジック） ----
            # 達成目標以上→時給+100円、維持目標以上達成未満→変動なし、維持目標未満→時給-100円
            if info["monthly_salary"] is None and r["target_achieve"] > 0 and r["target_maintain"] > 0:
                current_wage = info["hourly_wage"]
                if r["sales"] >= r["target_achieve"]:
                    r["confirmed_wage"] = current_wage + 100
                    r["confirmed_wage_change"] = "+100円（達成）"
                elif r["sales"] >= r["target_maintain"]:
                    r["confirmed_wage"] = current_wage
                    r["confirmed_wage_change"] = "変動なし（維持）"
                else:
                    r["confirmed_wage"] = current_wage - 100
                    r["confirmed_wage_change"] = "-100円（未達）"
            else:
                r["confirmed_wage"] = None
                r["confirmed_wage_change"] = None

            # ---- インセンティブ目標 ----
            # ロジック: (基本給 + 残業手当 + 非課税通勤手当) × 売上比重
            if info["monthly_salary"] is not None:
                r["incentive_status"] = "対象外（月給制）"
            else:
                wage = info["hourly_wage"]
                if wage < inc_min or wage > inc_max:
                    r["incentive_status"] = "対象外"
                else:
                    prior_work_days = prior_work_days_map.get(sid, 0)
                    rest_days = prior_business_days - prior_work_days

                    rate = None
                    week_label = None
                    for max_rest, rate_val, label in INCENTIVE_RATE_TABLE:
                        if rest_days <= max_rest:
                            rate = rate_val
                            week_label = label
                            break

                    payroll = prior_payroll_map.get(sid)

                    if rate is None:
                        r["incentive_status"] = "対象外（出勤実績不足）"
                    elif payroll is None:
                        r["incentive_status"] = "対象外（給与データ未登録）"
                    else:
                        payroll_total = (payroll["base_salary"]
                                          + payroll["overtime_allowance"]
                                          + payroll["commute_allowance"])
                        incentive_target = int(payroll_total * rate)

                        r["incentive_target"] = incentive_target
                        r["incentive_status"] = "ok"
                        r["incentive_detail"] = {
                            "prior_month": prior_month[:7],
                            "prior_work_days": prior_work_days,
                            "prior_business_days": prior_business_days,
                            "rest_days": rest_days,
                            "rate": rate,
                            "base_salary": payroll["base_salary"],
                            "overtime_allowance": payroll["overtime_allowance"],
                            "commute_allowance": payroll["commute_allowance"],
                            "payroll_total": payroll_total
                        }

                        # ---- 支給金額（インセンティブ還元額） ----
                        # ロジック: (目標値 - 2ヶ月前売上実績) × 還元率
                        prior_sales_actual = prior_sales_actual_map.get(sid, 0)
                        excess_amount = incentive_target - prior_sales_actual

                        # 時給帯に一致する行だけに絞る（ちょうどその時給の人のみ対象）
                        band_rows = [row for row in payout_rate_rows if row["wage_band"] == wage]

                        payout_rate = None
                        matched_segment = None
                        if payout_mode == "week":
                            for row in band_rows:
                                if row["segment_label"] == week_label:
                                    payout_rate = row["rate"]
                                    matched_segment = row["segment_label"]
                                    break
                        else:  # achieve_rate モード
                            if incentive_target > 0:
                                achieve_pct = round(prior_sales_actual / incentive_target * 100, 2)
                            else:
                                achieve_pct = 0
                            for row in band_rows:
                                smin = row.get("segment_min")
                                smax = row.get("segment_max")
                                if smin is not None and smax is not None and smin <= achieve_pct <= smax:
                                    payout_rate = row["rate"]
                                    matched_segment = row["segment_label"]
                                    break

                        if excess_amount <= 0:
                            r["incentive_payout"] = 0
                            r["incentive_payout_status"] = "対象外（超過なし）"
                        elif not band_rows:
                            r["incentive_payout"] = None
                            r["incentive_payout_status"] = "対象外（時給帯未登録）"
                        elif payout_rate is None:
                            r["incentive_payout"] = None
                            r["incentive_payout_status"] = "対象外（区分不一致）"
                        else:
                            payout = int(excess_amount * payout_rate)
                            r["incentive_payout"] = payout
                            r["incentive_payout_status"] = "ok"
                            r["incentive_detail"]["prior_sales_actual"] = prior_sales_actual
                            r["incentive_detail"]["excess_amount"] = excess_amount
                            r["incentive_detail"]["payout_rate"] = payout_rate
                            r["incentive_detail"]["payout_mode"] = payout_mode
                            r["incentive_detail"]["matched_segment"] = matched_segment

        return {"status": "ok", "data": list(results.values())}

    @app.route("/staff/summary")
    def staff_summary():
        try:
            month = request.args.get("month")
            result = build_staff_summary(month)
            status_code = 400 if "error" in result and result.get("error") == "monthパラメータが必要です" else 200
            return jsonify(result), status_code
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/appointments_json", methods=["POST"])
    def upload_appointments_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400

            # target_monthを設定
            for r in records:
                r["target_month"] = r["appointment_date"][:7] + "-01" if r.get("appointment_date") else None

            # アップロードデータに含まれる月の一覧
            months_in_data = sorted(set(r["target_month"] for r in records if r.get("target_month")))

            # まずupsertで全件保存（既存は上書き、新規は追加）
            supabase_staff.table("appointments").upsert(records, on_conflict="appointment_id").execute()

            # アポリストに存在しないIDを一括削除
            # 対象月に存在する全IDをDBから1回で取得
            uploaded_ids = set(r["appointment_id"] for r in records if r.get("appointment_id"))
            for month in months_in_data:
                existing_res = supabase_staff.table("appointments")\
                    .select("appointment_id")\
                    .eq("target_month", month)\
                    .execute()
                existing_ids = set(r["appointment_id"] for r in existing_res.data)
                ids_to_delete = existing_ids - uploaded_ids
                if ids_to_delete:
                    supabase_staff.table("appointments")\
                        .delete()\
                        .in_("appointment_id", list(ids_to_delete))\
                        .execute()

            return jsonify({"status": "ok", "count": len(records), "months": months_in_data})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/upload/productivity_json", methods=["POST"])
    def upload_productivity_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["call_date"][:7] + "-01" if r.get("call_date") else None
            supabase_staff.table("productivity").upsert(records, on_conflict="staff_id,call_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
    @app.route("/staff/upload/attendance_json", methods=["POST"])
    def upload_attendance_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400
            for r in records:
                r["target_month"] = r["work_date"][:7] + "-01" if r.get("work_date") else None
            supabase_staff.table("attendance").upsert(records, on_conflict="staff_id,work_date").execute()
            return jsonify({"status": "ok", "count": len(records)})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
                
    @app.route("/staff/upload/fb_bulk_json", methods=["POST"])
    @admin_required
    def upload_fb_bulk_json():
        try:
            data = request.get_json()
            records = data.get("records", [])
            if not records:
                return jsonify({"error": "データがありません"}), 400

            # campaign_nameからcampaign_idを解決
            campaigns_res = supabase_staff.table("fb_campaigns").select("id,name").execute()
            name_to_id = {c["name"]: c["id"] for c in campaigns_res.data}

            rows = []
            errors = []
            for r in records:
                sid = str(r.get("staff_id", "")).strip()
                campaign_name = str(r.get("campaign_name", "")).strip()
                amount = r.get("amount", 0)

                if not sid or not campaign_name:
                    continue

                campaign_id = name_to_id.get(campaign_name)
                if not campaign_id:
                    errors.append(f"キャンペーン名が見つかりません: {campaign_name}")
                    continue

                sid = B_TO_D.get(sid, sid)
                rows.append({
                    "campaign_id": campaign_id,
                    "staff_id": sid,
                    "amount": int(str(amount).replace(",", "")) if amount else 0
                })

            if rows:
                # 既存データを削除してから再挿入（同一キャンペーンの上書き）
                campaign_ids = list(set(r["campaign_id"] for r in rows))
                for cid in campaign_ids:
                    supabase_staff.table("fb_bulk_amounts").delete().eq("campaign_id", cid).execute()
                supabase_staff.table("fb_bulk_amounts").insert(rows).execute()

            return jsonify({
                "status": "ok",
                "count": len(rows),
                "errors": errors
            })
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500

    @app.route("/staff/daily_calendar")
    def staff_daily_calendar():
        """
        個人ページのカレンダー表示用。
        指定スタッフ・指定月の、日付ごとのアポ金額・キャンセル金額・個別案件一覧を返す。
        """
        try:
            staff_id = request.args.get("staff_id")
            month = request.args.get("month")
            if not staff_id or not month:
                return jsonify({"error": "staff_idとmonthが必要です"}), 400

            target_month = month + "-01"

            apo_res = supabase_staff.table("appointments")\
                .select("*").eq("target_month", target_month).execute()
            apo_rows = apo_res.data

            daily = {}
            for row in apo_rows:
                sid = B_TO_D.get(row["staff_id"], row["staff_id"])
                if sid != staff_id:
                    continue
                day = row.get("appointment_date")
                if not day:
                    continue
                if day not in daily:
                    daily[day] = {"apo_amount": 0, "cxl_amount": 0, "items": []}
                cancel = str(row.get("cancel_date") or "")
                amount = row.get("achievement_amount", 0)
                project_name = row.get("project_name") or "（案件名未登録）"
                is_cancel = bool(cancel and cancel not in ["None", ""])
                if is_cancel:
                    daily[day]["cxl_amount"] += amount
                else:
                    daily[day]["apo_amount"] += amount
                daily[day]["items"].append({
                    "project_name": project_name,
                    "amount": amount,
                    "is_cancel": is_cancel
                })

            result = []
            for day, vals in sorted(daily.items()):
                result.append({
                    "date": day,
                    "apo_amount": vals["apo_amount"],
                    "cxl_amount": vals["cxl_amount"],
                    "net_amount": vals["apo_amount"] - vals["cxl_amount"],
                    "items": vals["items"]
                })

            return jsonify({"status": "ok", "data": result})
        except Exception as e:
            import traceback
            return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500
